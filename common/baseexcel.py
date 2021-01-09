import openpyxl
from common.baselogger import logger


class BaseExcel:
    """
    excel操作的基类
    """

    @staticmethod
    def write_excel(file_path, sheet_name, row, column, data):
        """
        写入excel
        """
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb[sheet_name]
        except Exception as e:
            logger.error("打开excel失败")
            raise e
        cell = sheet.cell(row, column)
        cell.value = data
        wb.save(file_path)
        logger.info("写入excel {}成功，数据：{}".format(sheet_name, data))

    def __init__(self, file_path):
        self.wb = openpyxl.load_workbook(file_path)
        self.sheet = None

    def load_sheet(self, sheet_name):
        try:
            self.sheet = self.wb[sheet_name]
        except Exception as e:
            logger.exception("加载excel表单失败")
            raise e
        logger.info("加载excel表单成功")
        return True

    def get_title(self, sheet_name):
        """
        获取第一行的内容
        :return:
        """
        self.load_sheet(sheet_name)
        title = []
        row = self.sheet[1]
        for cell in row:
            title.append(cell.value)
        return title

    def read_excel(self, sheet_name):
        """
        读取excel所有内容
        :return:
        """
        title = self.get_title(sheet_name)
        rows = list(self.sheet.rows)
        value_list = []
        for row in rows[1:]:
            li = []
            for cell in row:
                li.append(cell.value)
            d = dict(zip(title, li))
            value_list.append(d)
        return value_list

    def get_value_by_id(self, sheet_name, num):
        """
        根据case_id来获取对应行的内容
        :param sheet_name: 表单名称
        :param num: case_id
        :return:
        """
        self.load_sheet(sheet_name)
        row = self.sheet[num + 1]
        value_list = []
        for cell in row:
            value_list.append(cell.value)
        return value_list


# if __name__ == '__main__':
#     excel = BaseExcel('E:\桌面文件\cases.xlsx', '注册接口')
#     res = excel.get_id_value(2)
#     print(res)
